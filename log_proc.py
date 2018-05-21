import json
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)

with open('./log_original.json') as fp:
    original_data = json.load(fp)
with open('./log_pruned.json') as fp:
    pruned_data = json.load(fp)

bandwidth0 = np.power(32, 2) * 3

###############################
# original
###############################

layer_cfg = ['', '']
acc = ['acc_original', '']
# delta_t = ['time_original', 0]
delta_ts = 0
# delta_t_computations = ['', '']
bandwidth = ['bandwidth_original', bandwidth0]
# all_conv_computations = ['', '']
for data in original_data:
    layer_cfg.append(data['layer_cfg'])
    acc.append(data['acc'])
    # delta_t.append(data['delta_t'])
    delta_ts += np.array(data['delta_ts'])
    # delta_t_computations.append(data['delta_t_computations'])
    bandwidth.append(data['bandwidth'])
    # all_conv_computations.append(data['all_conv_computations'])
delta_ts = ['time_original', 0]+list(delta_ts / len(original_data))

wb = Workbook()
# ws = wb.create_sheet("original", 0)
ws = wb.active
ws.append(layer_cfg)
ws.append(acc)
# ws.append(delta_t)
ws.append(delta_ts)
# ws.append(delta_t_computations)
ws.append(bandwidth)
# ws.append(all_conv_computations)

###############################
# pruned
###############################

layer_cfg = ['', '']
acc = ['acc_pruned', '']
# delta_t = ['time_pruned', 0]
delta_ts = 0
# delta_t_computations = ['', '']
bandwidth = ['bandwidth_pruned', bandwidth0]
# all_conv_computations = ['', '']
for data in pruned_data:
    layer_cfg.append(data['layer_cfg'])
    acc.append(data['acc'])
    # delta_t.append(data['delta_t'])
    delta_ts += np.array(data['delta_ts'])
    # delta_t_computations.append(data['delta_t_computations'])
    bandwidth.append(data['bandwidth'])
    # all_conv_computations.append(data['all_conv_computations'])
delta_ts = ['time_pruned', 0]+list(delta_ts / len(original_data))

# ws = wb.create_sheet("pruned", 0)
for i in range(3):
    ws.append(list())
ws.append(layer_cfg)
ws.append(acc)
# ws.append(delta_t)
ws.append(delta_ts)
# ws.append(delta_t_computations)
ws.append(bandwidth)
# ws.append(all_conv_computations)

for i in range(3):
    ws.append(list())
for index in range(len(layer_cfg)):
    if isinstance(layer_cfg[index], int):
        layer_cfg[index] = 'conv'
    elif layer_cfg[index] == '' and index == 1:
        layer_cfg[index] = 'og'
    elif layer_cfg[index] == '':
        pass
    else:
        layer_cfg[index] = 'pool'

ws.append(layer_cfg)

# draw chart
time_original = Reference(ws, min_col=1, min_row=3, max_col=20)
time_pruned = Reference(ws, min_col=1, min_row=10, max_col=20)
bandwidth_original = Reference(ws, min_col=1, min_row=4, max_col=20)
bandwidth_pruned = Reference(ws, min_col=1, min_row=11, max_col=20)
# line chart for time
c1 = LineChart()
c1.add_data(time_original, titles_from_data=True, from_rows=True)
c1.add_data(time_pruned, titles_from_data=True, from_rows=True)

cats = Reference(ws, min_col=2, min_row=15, max_col=20)
c1.set_categories(cats)

c1.x_axis.title = 'layers'
c1.y_axis.title = 'time elapsed (s)'
c1.y_axis.majorGridlines = None
c1.title = 'bandwidth/time'

# bar chart for bandwidth
c2 = BarChart()
c2.add_data(bandwidth_original, titles_from_data=True, from_rows=True)
c2.add_data(bandwidth_pruned, titles_from_data=True, from_rows=True)
c2.y_axis.axId = 200  # set axid other than 100
c2.y_axis.title = 'data volume (B)'

# Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
c1.y_axis.crosses = "max"
c1 += c2

ws.add_chart(c1, "D4")

wb.save('log.xlsx')
